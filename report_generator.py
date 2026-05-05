#!/usr/bin/env python3
"""
report_generator.py  —  Phase 2: Report Generator
==================================================
Takes a Nessus .audit file + one or more Nessus compliance CSV exports
and produces a filled Excel report.

  Single CSV  → one sheet named by host IP
  Multiple CSV / folder → one sheet per host (no Consolidated sheet)

Default Value is sourced ONLY from the .audit file.
Observed Value and Status are sourced ONLY from the CSV files.
Blank beats N/A — if a value is not found it is left empty.

Usage:
    # Single scan
    python report_generator.py --audit CIS_L1.audit --csv scan.csv

    # Folder of CSVs (one tab per host)
    python report_generator.py --audit CIS_L1.audit --csv ./results/

    # Individual files
    python report_generator.py --audit CIS_L1.audit --csv h1.csv h2.csv h3.csv
"""

import re
import sys
import csv
import math
import glob
import argparse
from pathlib import Path
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
#  ROW HEIGHT
# ─────────────────────────────────────────────────────────────────────────────

_FONT_PT  = 9
_LINE_H   = _FONT_PT * 1.35
_CHARS_CW = 1.05


def _line_count(text: str, col_w: float) -> float:
    if not text:
        return 1.0
    cpl = max(1, col_w * _CHARS_CW)
    n   = 0.0
    for para in str(text).split("\n"):
        p = para.strip()
        n += 0.4 if not p else math.ceil(len(p) / cpl)
    return max(1.0, n)


def _row_height(pairs: list) -> float:
    mx = max((_line_count(t, w) for t, w in pairs), default=1.0)
    return max(18, min(mx * _LINE_H + 4, 400))


# ─────────────────────────────────────────────────────────────────────────────
#  AUDIT FILE PARSER  (reused from audit_template.py — same logic)
# ─────────────────────────────────────────────────────────────────────────────

_BLOCK_RE = re.compile(
    r'<(custom_item|report[^>]*)>(.*?)</(custom_item|report)>',
    re.DOTALL | re.IGNORECASE,
)
_KV_RE = re.compile(
    r'^\s*([\w]+)\s*:\s*(?:"((?:[^"\\]|\\.)*?)"|([^\n\r]+))',
    re.MULTILINE,
)
_BENCH_PATS = [
    re.compile(r'^\d+\.\d+'),
    re.compile(r'^[A-Z][A-Z0-9]{1,8}-\d{2}-\d{4,}'),
    re.compile(r'^[A-Z]{2,10}-\d{4,}'),
    re.compile(r'^V-\d{4,}'),
]
_VAR_DEFAULTS = {
    "@PASSWORD_HISTORY@":        "24 or more passwords",
    "@MAXIMUM_PASSWORD_AGE@":    "1–365 days",
    "@MINIMUM_PASSWORD_AGE@":    "1 or more day",
    "@MINIMUM_PASSWORD_LENGTH@": "14 or more characters",
    "@LOCKOUT_DURATION@":        "15 or more minutes",
    "@LOCKOUT_THRESHOLD@":       "1–5 invalid attempts",
    "@LOCKOUT_RESET@":           "15 or more minutes",
}


def _is_check(desc: str) -> bool:
    return any(p.match(desc.strip()) for p in _BENCH_PATS)


def _parse_block(block: str) -> dict:
    item = {}
    for m in _KV_RE.finditer(block):
        k = m.group(1).lower().strip()
        v = (m.group(2) if m.group(2) is not None else m.group(3) or "").strip()
        v = v.replace('\\"', '"').replace('\\n', '\n').replace('\\t', '\t')
        if k not in item:
            item[k] = v
    return item


def _clean(text: str) -> str:
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _default_value(value_data: str) -> str:
    if not value_data:
        return ""
    for ph, human in _VAR_DEFAULTS.items():
        if ph in value_data:
            return human
    val = re.sub(
        r'\[(\d+)\.\.(MAX|\d+)\]',
        lambda m: (f"{m.group(1)} or more" if m.group(2) == "MAX"
                   else f"{m.group(1)} to {m.group(2)}"),
        value_data,
    )
    val = val.replace(' || ', ' or ').strip('"').strip()
    return val[:150] if len(val) > 150 else val


def _bench_name_from_file(text: str, filepath: str) -> str:
    m = re.search(r'<display_name>(.*?)</display_name>', text, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r'#\s*description\s*:\s*This .audit is designed against the (.+)', text)
    if m:
        return m.group(1).strip()
    return Path(filepath).stem.replace('_', ' ')


def _sort_key(benchmark: str):
    m = re.match(r'^([\d.]+)', benchmark.strip())
    if m:
        return [int(p) if p.isdigit() else p for p in m.group(1).split('.')]
    m2 = re.match(r'^([A-Z0-9-]+)', benchmark.strip())
    return [m2.group(1)] if m2 else [benchmark]


def parse_audit(filepaths: list) -> tuple:
    checks    = []
    seen      = set()
    bench_name = ""
    for fp in filepaths:
        text = Path(fp).read_text(encoding="utf-8", errors="replace")
        if not bench_name:
            bench_name = _bench_name_from_file(text, fp)
        cnt = 0
        for m in _BLOCK_RE.finditer(text):
            item = _parse_block(m.group(2))
            desc = item.get("description", "").strip()
            if not desc or not _is_check(desc) or desc in seen:
                continue
            seen.add(desc)
            checks.append({
                "benchmark":      desc,
                "description":    _clean(item.get("info", "")),
                "default_value":  _default_value(item.get("value_data", "")),
                "recommendation": _clean(item.get("solution", "")),
            })
            cnt += 1
        print(f"  {Path(fp).name}: {cnt} checks")
    checks.sort(key=lambda c: _sort_key(c["benchmark"]))
    return checks, bench_name or "Checks"


# ─────────────────────────────────────────────────────────────────────────────
#  CSV PARSER
# ─────────────────────────────────────────────────────────────────────────────

# Generic plugin names used in vuln-scan CSV exports
_GENERIC_PLUGINS = {
    "unix compliance checks",
    "windows compliance checks",
    "database compliance checks",
    "compliance checks",
    "cisco compliance checks",
    "vmware compliance checks",
}

# "check name"" : [STATUS]  — the extra " is a CSV quote-doubling artefact
_DESC_CHECK_RE = re.compile(
    r'^\s*"([^"]+)"\s*"?\s*:\s*\[?(PASSED|FAILED|WARNING|ERROR)\]?',
    re.IGNORECASE,
)


def _norm_id(s: str) -> str:
    """Extract the leading benchmark ID for matching."""
    m = re.match(r'^([\d]+(?:\.[\d]+)*)', s.strip())
    if m:
        return m.group(1)
    m2 = re.match(r'^([A-Z][A-Z0-9]{1,8}-\d{2}-\d{4,})', s.strip())
    if m2:
        return m2.group(1)
    return re.sub(r'[^a-z0-9]', '', s.lower())[:30]


def _norm_status(raw: str) -> str:
    r = (raw or "").strip().upper()
    if r in ("PASSED", "PASS"):   return "PASSED"
    if r in ("FAILED", "FAIL"):   return "FAILED"
    if r in ("WARNING", "WARN"):  return "WARNING"
    return r or ""


def _extract_from_description(desc: str) -> tuple:
    """
    Parse a vuln-scan CSV Description field.
    Returns (check_name, status, observed_value)

    Description structure:
      "check name"" : [STATUS]
      ...
      Policy Value:
        <expected regex/config>
      Actual Value:
        <what was found on the system>
        - Audit Result:
          ** PASS / FAIL **
    """
    m = _DESC_CHECK_RE.match(desc)
    if not m:
        return None, None, None

    check_name = m.group(1).strip()
    status     = m.group(2).upper()

    # Actual Value → from "Actual Value:\n" to end of string
    av_m = re.search(r'Actual\s+Value:\s*\n(.*)', desc, re.DOTALL | re.IGNORECASE)
    if av_m:
        raw = av_m.group(1).strip()
        # Clean CSV quote-doubling artefacts
        raw = raw.replace('"""', '"').replace('""', '"')
        observed = raw[:800]
    else:
        observed = ""

    return check_name, status, observed


def _host_label(csv_path: Path, rows: list) -> str:
    """Return host IP if found in rows, otherwise clean filename."""
    for row in rows[:20]:
        h = (row.get('Host') or row.get('host') or "").strip()
        if re.match(r'\d{1,3}\.\d{1,3}', h):
            return h
        if h:
            return h[:31]
    stem = re.sub(r'^(scan|nessus|result|compliance)[_-]', '',
                  csv_path.stem, flags=re.I)
    return stem[:31]


def parse_csv(csv_path: Path) -> tuple:
    """
    Parse one Nessus compliance CSV.
    Returns (host_label, { norm_id: {status, observed_value} })

    Uses default csv.DictReader (csv.excel dialect) — do NOT use csv.Sniffer
    as it breaks multi-line quoted fields which are common in Nessus CSVs.
    """
    results  = {}
    rows_raw = []

    try:
        with open(csv_path, newline="", encoding="utf-8-sig",
                  errors="replace") as f:
            reader = csv.DictReader(f)   # default dialect handles multi-line fields
            hdrs   = reader.fieldnames or []

            # Detect column names
            name_col   = next((c for c in ("Name", "Plugin Name") if c in hdrs), "")
            status_col = next((c for c in ("Risk", "Status", "Result") if c in hdrs), "")
            host_col   = next((c for c in ("Host", "IP Address") if c in hdrs), "")

            if not name_col or not status_col:
                print(f"  ⚠  {csv_path.name}: cannot find Name/Status columns")
                return csv_path.stem, {}

            for row in reader:
                name   = (row.get(name_col) or "").strip()
                status = _norm_status(row.get(status_col) or "")
                host   = (row.get(host_col) or "").strip()

                if not status:
                    continue

                # Detect format: generic plugin name means vuln-scan CSV
                if name.lower() in _GENERIC_PLUGINS:
                    desc = (row.get("Description") or "").strip()
                    check_name, status, observed = _extract_from_description(desc)
                    if not check_name:
                        continue
                else:
                    # Standard compliance CSV — check name is in Name column
                    check_name = name
                    # Plugin Output is empty in many exports; Description has the data
                    plugin_out = (row.get("Plugin Output") or "").strip()
                    if plugin_out:
                        observed = plugin_out[:800]
                    else:
                        # Extract from Description if available
                        desc = (row.get("Description") or "").strip()
                        _, _, observed = _extract_from_description(desc)
                        observed = observed or ""

                if not check_name or not _is_check(check_name):
                    continue

                rows_raw.append(row)
                key = _norm_id(check_name)

                if key not in results:
                    results[key] = {
                        "status":         status,
                        "observed_value": observed,
                        "host":           host,
                    }
                else:
                    # Worst-case: FAILED > WARNING > PASSED
                    existing = results[key]["status"]
                    if status == "FAILED" or (existing != "FAILED"
                                              and status == "WARNING"):
                        results[key]["status"] = status
                    if observed and observed not in results[key]["observed_value"]:
                        results[key]["observed_value"] = (
                            results[key]["observed_value"] + "\n" + observed
                        ).strip()

    except Exception as e:
        print(f"  ⚠  {csv_path.name}: {e}")
        return csv_path.stem, {}

    label = _host_label(csv_path, rows_raw)
    return label, results


def expand_inputs(inputs: list) -> list:
    """Expand file paths and folders to a flat list of .csv Paths."""
    files = []
    for inp in inputs:
        p = Path(inp)
        if p.is_dir():
            found = sorted(p.glob("*.csv"))
            if not found:
                print(f"  ⚠  No .csv files in: {inp}")
            files.extend(found)
        elif p.is_file():
            files.append(p)
        else:
            matched = sorted(Path(x) for x in glob.glob(inp))
            if matched:
                files.extend(matched)
            else:
                print(f"  ⚠  Not found: {inp}")
    return files


def merge(base_checks: list, csv_results: dict) -> tuple:
    """Fill base_checks with status/observed_value from csv_results."""
    checks  = deepcopy(base_checks)
    matched = 0
    for chk in checks:
        key  = _norm_id(chk["benchmark"])
        data = csv_results.get(key)
        if not data:
            # Fuzzy: prefix overlap
            for k, v in csv_results.items():
                if k.startswith(key) or key.startswith(k):
                    data = v
                    break
        if data:
            chk["status"]         = data["status"]
            chk["observed_value"] = data["observed_value"]
            matched += 1
        else:
            chk["status"]         = ""
            chk["observed_value"] = ""
    return checks, matched


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL STYLES
# ─────────────────────────────────────────────────────────────────────────────

C_HDR_BG  = "1F3864";  C_HDR_FG  = "FFFFFF"
C_CAT_BG  = "2E75B6";  C_CAT_FG  = "FFFFFF"
C_PASS_BG = "EBF5E1";  C_PASS_FG = "375623"
C_FAIL_BG = "FDE9E8";  C_FAIL_FG = "9C0006"
C_WARN_BG = "FFF3CD";  C_WARN_FG = "7F5200"
C_ALT     = "EBF3FB";  C_WHITE   = "FFFFFF"
C_BORDER  = "8EA9C1"

_S   = Side(style="thin", color=C_BORDER)
_BRD = Border(left=_S, right=_S, top=_S, bottom=_S)

COLUMNS = [
    ("S.NO",           7,   False, "center"),
    ("Benchmark",      52,  True,  "left"),
    ("Description",    70,  True,  "left"),
    ("Status",         14,  False, "center"),
    ("Default Value",  28,  True,  "left"),
    ("Observed Value", 32,  True,  "left"),
    ("Recommendation", 62,  True,  "left"),
]


def _status_style(s: str):
    s = s.upper()
    if s == "PASSED":  return C_PASS_BG, C_PASS_FG
    if s == "FAILED":  return C_FAIL_BG, C_FAIL_FG
    if s == "WARNING": return C_WARN_BG, C_WARN_FG
    return C_WHITE, "595959"


def _hdr(cell, text: str):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color=C_HDR_FG, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _BRD


def _dat(cell, value, bg: str, fg: str, align: str, wrap: bool, bold=False):
    cell.value     = value
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", size=_FONT_PT, color=fg, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="top",
                               wrap_text=wrap)
    cell.border    = _BRD


def write_checks_sheet(ws, checks: list):
    ws.row_dimensions[1].height = 30
    for ci, (label, width, _, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        _hdr(ws.cell(row=1, column=ci), label)
    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    for ri, chk in enumerate(checks, start=1):
        row    = ri + 1
        status = chk.get("status", "")
        if status:
            row_bg, sfg = _status_style(status)
        else:
            row_bg, sfg = (C_ALT if ri % 2 == 0 else C_WHITE), "595959"

        vals = [
            (ri,                            "center", False, False),
            (chk["benchmark"],              "left",   True,  False),
            (chk["description"],            "left",   True,  False),
            (status,                        "center", False, True),
            (chk["default_value"],          "left",   True,  False),
            (chk.get("observed_value", ""), "left",   True,  False),
            (chk["recommendation"],         "left",   True,  False),
        ]
        for ci, (val, align, wrap, bold) in enumerate(vals, start=1):
            label = COLUMNS[ci-1][0]
            fg    = sfg if label == "Status" else "000000"
            _dat(ws.cell(row=row, column=ci),
                 val if ci > 1 else ri, row_bg, fg, align, wrap, bold)

        wrapped = [(str(v), COLUMNS[ci-1][1])
                   for ci, (v, _, wrap, _) in enumerate(vals, start=1)
                   if wrap]
        ws.row_dimensions[row].height = _row_height(wrapped)

    # Tab colour by worst status
    statuses = {chk.get("status","").upper() for chk in checks if chk.get("status")}
    if "FAILED"  in statuses: ws.sheet_properties.tabColor = "C2212E"
    elif "WARNING" in statuses: ws.sheet_properties.tabColor = "F18C43"
    elif "PASSED" in statuses: ws.sheet_properties.tabColor = "527421"


def write_summary(wb: Workbook, bench_name: str,
                  host_data: list, total_checks: int):
    """
    host_data = [(host_label, checks_list), ...]
    Single host: show overall stats only.
    Multiple hosts: show per-host table.
    """
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_properties.tabColor = "1F3864"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    span = "A1:E1" if len(host_data) > 1 else "A1:B1"
    ws.merge_cells(span)
    t           = ws.cell(row=1, column=1, value=bench_name)
    t.font      = Font(name="Arial", bold=True, size=13, color=C_HDR_FG)
    t.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    if len(host_data) == 1:
        # Single host — vertical stats
        checks = host_data[0][1]
        p = sum(1 for c in checks if c.get("status") == "PASSED")
        f = sum(1 for c in checks if c.get("status") == "FAILED")
        w = sum(1 for c in checks if c.get("status") == "WARNING")
        n = total_checks - p - f - w
        for i, (lbl, val, fg, bg) in enumerate([
            ("Total Checks",    total_checks, C_HDR_BG, C_HDR_FG),
            ("✔  Passed",       p,            "375623", C_PASS_BG),
            ("✘  Failed",       f,            "9C0006", C_FAIL_BG),
            ("⚠  Warning",      w,            "7F5200", C_WARN_BG),
            ("—  Not Evaluated",n,            "595959", "F2F2F2"),
        ], start=3):
            for ci, (v, al) in enumerate([(lbl, "left"), (val, "center")], start=1):
                c = ws.cell(row=i, column=ci, value=v)
                c.font      = Font(name="Arial", bold=True, size=10, color=fg)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal=al, vertical="center",
                                        indent=1 if al == "left" else 0)
                c.border    = _BRD
            ws.row_dimensions[i].height = 22
    else:
        # Multi-host — table
        for ci, h in enumerate(["Host", "Total", "Passed", "Failed", "Not Eval"],
                               start=1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font      = Font(name="Arial", bold=True, color=C_CAT_FG, size=10)
            c.fill      = PatternFill("solid", fgColor=C_CAT_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _BRD
        ws.row_dimensions[3].height = 22

        for ri, (label, checks) in enumerate(host_data, start=4):
            p  = sum(1 for c in checks if c.get("status") == "PASSED")
            f  = sum(1 for c in checks if c.get("status") == "FAILED")
            n  = total_checks - p - f
            alt = ri % 2 == 0
            bg  = C_ALT if alt else C_WHITE
            for ci, val in enumerate([label, total_checks, p, f, n], start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", size=9, bold=(ci > 1),
                              color=(C_PASS_FG if ci == 3 and p > 0 else
                                     C_FAIL_FG if ci == 4 and f > 0 else "000000"))
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(
                    horizontal="left" if ci == 1 else "center",
                    vertical="center", indent=1 if ci == 1 else 0)
                c.border    = _BRD
            ws.row_dimensions[ri].height = 18

        # Total row
        tr = 4 + len(host_data)
        all_p = sum(p for _, chks in host_data
                    for p in [sum(1 for c in chks if c.get("status")=="PASSED")])
        all_f = sum(f for _, chks in host_data
                    for f in [sum(1 for c in chks if c.get("status")=="FAILED")])
        for ci, (v, al) in enumerate(
                [("TOTAL", "left"), (total_checks, "center"),
                 (all_p, "center"), (all_f, "center"),
                 (total_checks - all_p - all_f, "center")], start=1):
            c = ws.cell(row=tr, column=ci, value=v)
            c.font      = Font(name="Arial", bold=True, color=C_HDR_FG, size=10)
            c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
            c.alignment = Alignment(horizontal=al, vertical="center",
                                    indent=1 if al == "left" else 0)
            c.border    = _BRD
        ws.row_dimensions[tr].height = 22


def _safe_tab(s: str) -> str:
    return re.sub(r'[:/\\?\[\]*]', '-', s)[:31]


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_report(base_checks: list, bench_name: str,
                 csv_paths: list, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    host_data = []
    used_tabs = {}

    print(f"\n[2] Parsing {len(csv_paths)} CSV file(s)…")
    for cp in csv_paths:
        label, results = parse_csv(cp)
        filled, matched = merge(base_checks, results)
        host_data.append((label, filled))
        print(f"  {cp.name:<42} host={label}  matched={matched}/{len(base_checks)}")

    print(f"\n[3] Building Excel…")
    for label, checks in host_data:
        safe = _safe_tab(label)
        if safe in used_tabs:
            used_tabs[safe] += 1
            safe = _safe_tab(f"{label}_{used_tabs[safe]}")
        else:
            used_tabs[safe] = 1

        ws = wb.create_sheet(title=safe)
        write_checks_sheet(ws, checks)

        p = sum(1 for c in checks if c.get("status") == "PASSED")
        f = sum(1 for c in checks if c.get("status") == "FAILED")
        print(f"  [tab] {safe:<32}  {p}✔  {f}✘")

    write_summary(wb, bench_name, host_data, len(base_checks))
    wb.save(output_path)

    print(f"\n[✓] Report saved → {output_path}")
    print(f"    Benchmark  : {bench_name}")
    print(f"    Checks     : {len(base_checks)}")
    if len(csv_paths) == 1:
        p = sum(1 for c in host_data[0][1] if c.get("status") == "PASSED")
        f = sum(1 for c in host_data[0][1] if c.get("status") == "FAILED")
        n = len(base_checks) - p - f
        print(f"    Results    : {p} PASSED  |  {f} FAILED  |  {n} Not Evaluated")
    else:
        print(f"    Tabs       : Summary + {len(csv_paths)} host sheet(s)")


# ─────────────────────────────────────────────────────────────────────────────
#  CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Phase 2 — Generate a filled compliance report from .audit + Nessus CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single scan
  python report_generator.py --audit CIS_L1.audit --csv scan.csv

  # Folder of CSVs → one tab per host
  python report_generator.py --audit CIS_L1.audit --csv ./results/

  # Explicit list
  python report_generator.py --audit CIS_L1.audit --csv h1.csv h2.csv h3.csv -o report.xlsx
""")
    ap.add_argument("--audit", nargs="+", required=True, metavar="FILE")
    ap.add_argument("--csv",   nargs="+", required=True, metavar="FILE_OR_FOLDER")
    ap.add_argument("-o", "--output", default="compliance_report.xlsx")
    args = ap.parse_args()

    missing = [f for f in args.audit if not Path(f).exists()]
    if missing:
        for f in missing: print(f"[ERROR] Not found: {f}", file=sys.stderr)
        sys.exit(1)

    csv_paths = expand_inputs(args.csv)
    if not csv_paths:
        print("[ERROR] No CSV files found.", file=sys.stderr)
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  {len(csv_paths)} CSV file(s)  →  "
          f"{'one tab per host' if len(csv_paths) > 1 else 'single sheet'}")
    print(f"{'='*60}")

    print(f"\n[1] Parsing {len(args.audit)} audit file(s)…")
    base_checks, bench_name = parse_audit(args.audit)
    print(f"    Total checks : {len(base_checks)}")
    print(f"    Benchmark    : {bench_name}")

    if not base_checks:
        print("[ERROR] No checks found in audit file.")
        sys.exit(1)

    build_report(base_checks, bench_name, csv_paths, args.output)


if __name__ == "__main__":
    main()
