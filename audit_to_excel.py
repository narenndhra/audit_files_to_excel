#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          Universal Nessus Audit → Excel  (Manual + Automatic modes)        ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Works with ANY .audit file: CIS Benchmarks, DISA STIGs, and all others    ║
║                                                                              ║
║  MANUAL MODE  — audit only, blank Status + Observed Value                   ║
║    python audit_to_excel.py --audit CIS_L1.audit                            ║
║                                                                              ║
║  AUTOMATIC MODE — single CSV                                                 ║
║    python audit_to_excel.py --audit CIS_L1.audit --csv scan.csv             ║
║                                                                              ║
║  MULTI-HOST — individual CSV files                                           ║
║    python audit_to_excel.py --audit CIS_L1.audit --csv h1.csv h2.csv h3.csv ║
║                                                                              ║
║  MULTI-HOST — folder of CSVs (all *.csv files inside)                       ║
║    python audit_to_excel.py --audit CIS_L1.audit --csv ./results/           ║
║                                                                              ║
║  Multiple audit files:                                                       ║
║    python audit_to_excel.py --audit L1.audit L2.audit --csv ./results/      ║
╚══════════════════════════════════════════════════════════════════════════════╝

Output for multi-host:
  Summary      — overall stats + per-host pass/fail counts
  Consolidated — worst-case per check across all hosts
  192.168.1.10 — individual results for host 1   (tab = IP or CSV filename)
  192.168.1.11 — individual results for host 2
  ...

Output for single CSV / manual:
  Summary      — overall stats
  <BenchmarkName> — all checks (blank in manual mode)
"""

import re
import sys
import csv
import math
import glob
import argparse
from pathlib import Path
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  ROW HEIGHT  (openpyxl never auto-sizes — we calculate explicitly)
# ══════════════════════════════════════════════════════════════════════════════

FONT_PT       = 9
LINE_H        = FONT_PT * 1.35
MIN_ROW_H     = 18
MAX_ROW_H     = 400
CHARS_PER_COL = 1.05


def _lines(text: str, col_w: float) -> float:
    if not text:
        return 1.0
    cpl = max(1, col_w * CHARS_PER_COL)
    tot = 0.0
    for para in str(text).split("\n"):
        p = para.strip()
        tot += 0.4 if not p else math.ceil(len(p) / cpl)
    return max(1.0, tot)


def row_height(pairs: list) -> float:
    mx = max((_lines(t, w) for t, w in pairs), default=1.0)
    return max(MIN_ROW_H, min(mx * LINE_H + 4, MAX_ROW_H))


# ══════════════════════════════════════════════════════════════════════════════
#  AUDIT FILE PARSER
# ══════════════════════════════════════════════════════════════════════════════

_BLOCK_RE = re.compile(
    r'<(custom_item|report[^>]*)>(.*?)</(custom_item|report)>',
    re.DOTALL | re.IGNORECASE,
)
_KV_RE = re.compile(
    r'^\s*([\w]+)\s*:\s*(?:"((?:[^"\\]|\\.)*?)"|([^\n\r]+))',
    re.MULTILINE,
)
_BENCH_PATS = [
    re.compile(r'^\d+\.\d+'),                          # CIS:  1.1.1 ...
    re.compile(r'^[A-Z][A-Z0-9]{1,8}-\d{2}-\d{4,}'),  # DISA: O19C-00-000200
    re.compile(r'^[A-Z]{2,10}-\d{4,}'),                # Other STIGs
    re.compile(r'^V-\d{4,}'),                          # Vuln-ID style
]
_VAR_MAP = {
    "@PASSWORD_HISTORY@":        "24 or more passwords",
    "@MAXIMUM_PASSWORD_AGE@":    "1 to 365 days",
    "@MINIMUM_PASSWORD_AGE@":    "1 or more day(s)",
    "@MINIMUM_PASSWORD_LENGTH@": "14 or more characters",
    "@LOCKOUT_DURATION@":        "15 or more minutes",
    "@LOCKOUT_THRESHOLD@":       "1 to 5 invalid attempts",
    "@LOCKOUT_RESET@":           "15 or more minutes",
}


def _is_check(d: str) -> bool:
    return any(p.match(d.strip()) for p in _BENCH_PATS)


def _parse_block(block: str) -> dict:
    item = {}
    for m in _KV_RE.finditer(block):
        k = m.group(1).lower().strip()
        v = (m.group(2) if m.group(2) is not None else m.group(3) or '').strip()
        v = v.replace('\\"', '"').replace('\\n', '\n').replace('\\t', '\t')
        if k not in item:
            item[k] = v
    return item


def _clean(t: str) -> str:
    return re.sub(r'\n{3,}', '\n\n', re.sub(r'[ \t]+', ' ', t)).strip()


def _defval(vd: str) -> str:
    if not vd:
        return ""
    for ph, hm in _VAR_MAP.items():
        if ph in vd:
            return hm
    v = re.sub(r'\[(\d+)\.\.(MAX|\d+)\]',
               lambda m: f"{m.group(1)} or more" if m.group(2) == "MAX"
                          else f"{m.group(1)} to {m.group(2)}", vd)
    return v.replace(' || ', ' or ').strip('"').strip()


def _bench_name(text: str, fp: str) -> str:
    m = re.search(r'<display_name>(.*?)</display_name>', text, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r'#\s*description\s*:\s*This .audit is designed against the (.+)', text)
    if m:
        return m.group(1).strip()
    return Path(fp).stem.replace('_', ' ')


def parse_audit_files(filepaths: list) -> tuple:
    """Returns (checks, benchmark_name). Each check is a fresh dict."""
    checks, seen, name = [], set(), ""
    for fp in filepaths:
        text = Path(fp).read_text(encoding="utf-8", errors="replace")
        if not name:
            name = _bench_name(text, fp)
        cnt = 0
        for m in _BLOCK_RE.finditer(text):
            item = _parse_block(m.group(2))
            d    = item.get("description", "").strip()
            if not d or not _is_check(d) or d in seen:
                continue
            seen.add(d)
            checks.append({
                "benchmark":      d,
                "description":    _clean(item.get("info", "")),
                "default_value":  _defval(item.get("value_data", "")),
                "recommendation": _clean(item.get("solution", "")),
            })
            cnt += 1
        print(f"    {Path(fp).name}: {cnt} checks")

    # Natural sort
    def _key(c):
        d = c["benchmark"]
        m = re.match(r'^([\d.]+)', d)
        if m:
            return [int(p) if p.isdigit() else p for p in m.group(1).split('.')]
        m2 = re.match(r'^([A-Z0-9-]+)', d)
        return [m2.group(1)] if m2 else [d]

    checks.sort(key=_key)
    return checks, name or "Checks"


# ══════════════════════════════════════════════════════════════════════════════
#  CSV PARSING
# ══════════════════════════════════════════════════════════════════════════════

_N_COLS = ("Name", "Plugin Name", "Check Name", "name")
_S_COLS = ("Risk", "Status", "Result", "risk", "status")
_O_COLS = ("Plugin Output", "Output", "Actual Value", "plugin_output")
_H_COLS = ("Host", "IP Address", "host", "ip_address")


def _fcol(hdrs: list, cands: tuple) -> str:
    return next((c for c in cands if c in hdrs), "")


def _actual(out: str) -> str:
    if not out:
        return ""
    m = re.search(r'Actual\s+Value\s*:\s*(.+?)(?:\n|Policy|$)', out, re.I | re.S)
    if m:
        return m.group(1).strip()
    for ln in out.strip().splitlines():
        ln = ln.strip()
        if ln and not re.match(r'(policy|expected|check)', ln, re.I):
            return ln[:200]
    return out.strip()[:200]


def _norm_status(r: str) -> str:
    r = r.strip().upper()
    if r in ("PASSED", "PASS"):   return "PASSED"
    if r in ("FAILED", "FAIL"):   return "FAILED"
    if r in ("WARNING", "WARN"):  return "WARNING"
    return r or "INFO"


def _norm_id(s: str) -> str:
    m = re.match(r'^([\d]+(?:\.[\d]+)*)', s.strip())
    if m:
        return m.group(1)
    m2 = re.match(r'^([A-Z][A-Z0-9]{1,8}-\d{2}-\d{4,})', s.strip())
    if m2:
        return m2.group(1)
    return re.sub(r'[^a-z0-9]', '', s.lower())[:30]


def _expand_csv_paths(inputs: list) -> list:
    """
    Expand a mixed list of file paths and/or folder paths into CSV file list.
    --csv ./folder/     → all *.csv inside that folder
    --csv h1.csv h2.csv → those files
    --csv ./folder/ extra.csv → combined
    """
    files = []
    for inp in inputs:
        p = Path(inp)
        if p.is_dir():
            found = sorted(p.glob("*.csv"))
            if not found:
                print(f"  ⚠  No .csv files found in folder: {inp}")
            files.extend(found)
        elif p.is_file():
            files.append(p)
        else:
            # Glob pattern support e.g. "./results/*.csv"
            matched = [Path(x) for x in glob.glob(inp)]
            if matched:
                files.extend(sorted(matched))
            else:
                print(f"  ⚠  Not found: {inp}")
    return files


def _extract_host_label(csv_path: Path, rows: list) -> str:
    """
    Return the best label for a CSV file's tab name.
    Priority: first IP found in rows → cleaned filename.
    """
    for row in rows[:10]:
        h = row.get("_host", "").strip()
        if h and re.match(r'\d{1,3}\.\d{1,3}', h):
            return h
        if h:
            return h[:31]
    # Fallback to filename (strip extension, shorten)
    stem = csv_path.stem
    # Remove common prefixes like "scan_", "nessus_"
    stem = re.sub(r'^(scan|nessus|result|compliance)[_-]', '', stem, flags=re.I)
    return stem[:31]


def parse_csv_file(csv_path: Path) -> tuple:
    """
    Parse one CSV.  Returns (host_label, {norm_id: {status, observed_value}})
    """
    rows_raw = []
    results  = {}
    try:
        with open(csv_path, newline="", encoding="utf-8-sig", errors="replace") as f:
            sample  = f.read(4096); f.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
            reader  = csv.DictReader(f, dialect=dialect)
            hdrs    = reader.fieldnames or []

            name_col   = _fcol(hdrs, _N_COLS)
            status_col = _fcol(hdrs, _S_COLS)
            output_col = _fcol(hdrs, _O_COLS)
            host_col   = _fcol(hdrs, _H_COLS)

            if not name_col or not status_col:
                print(f"  ⚠  {csv_path.name}: can't find Name/Status columns")
                return csv_path.stem, {}

            for row in reader:
                name = row.get(name_col, "").strip()
                if not name:
                    continue
                row["_host"] = row.get(host_col, "")
                rows_raw.append(row)

                key    = _norm_id(name)
                status = _norm_status(row.get(status_col, ""))
                obs    = _actual(row.get(output_col, ""))
                host   = row.get(host_col, "").strip()

                if key not in results:
                    results[key] = {"status": status, "observed_value": obs,
                                    "host": host}
                else:
                    # Worst-case if the same check appears more than once
                    existing = results[key]["status"]
                    if status == "FAILED" or (existing != "FAILED" and status == "WARNING"):
                        results[key]["status"] = status
                    if obs and obs not in results[key]["observed_value"]:
                        results[key]["observed_value"] += ("\n" + obs).lstrip()

    except Exception as e:
        print(f"  ⚠  {csv_path.name}: {e}")
        return csv_path.stem, {}

    label = _extract_host_label(csv_path, rows_raw)
    return label, results


def apply_csv_to_checks(base_checks: list, csv_results: dict) -> list:
    """
    Clone base_checks and fill status/observed_value from csv_results.
    Returns new list (does NOT mutate base_checks).
    """
    checks = deepcopy(base_checks)
    matched = 0
    for chk in checks:
        key  = _norm_id(chk["benchmark"])
        data = csv_results.get(key)
        if not data:
            for k, v in csv_results.items():
                if k.startswith(key) or key.startswith(k):
                    data = v
                    break
        if data:
            chk["status"]         = data["status"]
            chk["observed_value"] = data["observed_value"]
            matched += 1
        else:
            chk["status"]         = ""
            chk["observed_value"] = ""
    return checks, matched


def consolidate(host_checks_list: list, base_checks: list) -> list:
    """
    Worst-case consolidation across all hosts.
    For each check: FAILED if any host failed, else WARNING if any warned,
    else PASSED if all passed, else blank.
    Observed Value = unique values joined across all hosts.
    """
    from collections import defaultdict
    agg: dict = defaultdict(lambda: {"statuses": [], "obs": []})

    for checks in host_checks_list:
        for chk in checks:
            key = chk["benchmark"]
            if chk["status"]:
                agg[key]["statuses"].append(chk["status"])
            if chk["observed_value"]:
                agg[key]["obs"].append(chk["observed_value"])

    result = deepcopy(base_checks)
    for chk in result:
        key  = chk["benchmark"]
        data = agg.get(key, {})
        sts  = data.get("statuses", [])
        obs  = list(dict.fromkeys(data.get("obs", [])))

        if not sts:
            chk["status"]         = ""
            chk["observed_value"] = ""
        elif "FAILED"  in sts: chk["status"] = "FAILED"
        elif "WARNING" in sts: chk["status"] = "WARNING"
        elif all(s == "PASSED" for s in sts): chk["status"] = "PASSED"
        else: chk["status"] = sts[0]

        chk["observed_value"] = "  |  ".join(obs)[:500]

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL STYLES
# ══════════════════════════════════════════════════════════════════════════════

C_HDR_BG  = "1F3864";  C_HDR_FG  = "FFFFFF"
C_CAT_BG  = "2E75B6";  C_CAT_FG  = "FFFFFF"
C_PASS_BG = "EBF5E1";  C_PASS_FG = "375623"
C_FAIL_BG = "FDE9E8";  C_FAIL_FG = "9C0006"
C_WARN_BG = "FFF3CD";  C_WARN_FG = "7F5200"
C_ALT     = "EBF3FB";  C_WHITE   = "FFFFFF"
C_BORDER  = "8EA9C1"

_S   = Side(style="thin", color=C_BORDER)
_BRD = Border(left=_S, right=_S, top=_S, bottom=_S)

COLUMNS = [
    ("S.NO",           7,   False, "center"),
    ("Benchmark",      50,  True,  "left"),
    ("Description",    70,  True,  "left"),
    ("Status",         14,  False, "center"),
    ("Default Value",  26,  True,  "left"),
    ("Observed Value", 32,  True,  "left"),
    ("Recommendation", 62,  True,  "left"),
]


def _status_style(s: str):
    s = s.upper()
    if s == "PASSED":  return C_PASS_BG, C_PASS_FG
    if s == "FAILED":  return C_FAIL_BG, C_FAIL_FG
    if s == "WARNING": return C_WARN_BG, C_WARN_FG
    return C_WHITE, "595959"


def _tab_color(checks):
    statuses = {c.get("status", "").upper() for c in checks if c.get("status")}
    if "FAILED"  in statuses: return "C2212E"
    if "WARNING" in statuses: return "F18C43"
    if "PASSED"  in statuses: return "527421"
    return None


def _hdr(cell, text):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color=C_HDR_FG, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _BRD


def _dat(cell, value, bg, fg, align, wrap, bold=False):
    cell.value     = value
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", size=FONT_PT, color=fg, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border    = _BRD


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET WRITERS
# ══════════════════════════════════════════════════════════════════════════════

def _safe_title(s: str) -> str:
    return re.sub(r'[:/\\?\[\]*]', '-', s)[:31]


def write_checks_sheet(ws, checks: list, auto_mode: bool):
    ws.row_dimensions[1].height = 30
    for ci, (label, width, wrap, align) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        _hdr(ws.cell(row=1, column=ci), label)
    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    for ri, chk in enumerate(checks, start=1):
        row    = ri + 1
        status = chk.get("status", "")
        if auto_mode and status:
            row_bg, sfg = _status_style(status)
        else:
            row_bg, sfg = (C_ALT if ri % 2 == 0 else C_WHITE), "595959"

        vals = [
            (ri,                            "center", False, False),
            (chk["benchmark"],              "left",   True,  False),
            (chk["description"],            "left",   True,  False),
            (status,                        "center", False, True),
            (chk["default_value"],          "left",   True,  False),
            (chk.get("observed_value", ""), "left",   True,  False),
            (chk["recommendation"],         "left",   True,  False),
        ]
        for ci, (val, align, wrap, bold) in enumerate(vals, start=1):
            fg = sfg if COLUMNS[ci-1][0] == "Status" else "000000"
            _dat(ws.cell(row=row, column=ci),
                 val if ci > 1 else ri,
                 row_bg, fg, align, wrap, bold)

        wrapped = [(str(v), COLUMNS[ci-1][1])
                   for ci, (v, _, w, _) in enumerate(vals, start=1) if w]
        ws.row_dimensions[row].height = row_height(wrapped)

    tc = _tab_color(checks)
    if tc:
        ws.sheet_properties.tabColor = tc


def write_summary_sheet(wb: Workbook, bench_name: str,
                        host_data: list, base_count: int):
    """
    host_data = list of (host_label, checks_list)
    If single host, just show overall totals.
    If multi-host, show per-host breakdown table.
    """
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_properties.tabColor = "1F3864"

    multi = len(host_data) > 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    # Title
    span = "A1:E1" if multi else "A1:B1"
    ws.merge_cells(span)
    t = ws.cell(row=1, column=1, value=bench_name)
    t.font      = Font(name="Arial", bold=True, size=13, color=C_HDR_FG)
    t.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2" if multi else "A2:B2")
    mode = ("Automatic Mode  —  Multi-Host  (%d hosts)" % len(host_data)
            if multi else
            "Automatic Mode  (audit + CSV merged)" if host_data else
            "Manual Mode  (audit only — fill Status & Observed Value)")
    sub = ws.cell(row=2, column=1, value=mode)
    sub.font      = Font(name="Arial", size=9, italic=True, color="595959")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    if not host_data:
        # Manual mode — just total
        _summary_stat_row(ws, 4, "Total Checks", base_count, C_HDR_BG, C_HDR_FG)
        return

    if not multi:
        # Single host — simple stats
        checks = host_data[0][1]
        p = sum(1 for c in checks if c.get("status") == "PASSED")
        f = sum(1 for c in checks if c.get("status") == "FAILED")
        w = sum(1 for c in checks if c.get("status") == "WARNING")
        n = len(checks) - p - f - w
        for i, (lbl, val, fg, bg) in enumerate([
            ("Total Checks",    len(checks), C_HDR_BG, C_HDR_FG),
            ("✔  Passed",       p,           "375623", C_PASS_BG),
            ("✘  Failed",       f,           "9C0006", C_FAIL_BG),
            ("⚠  Warning",      w,           "7F5200", C_WARN_BG),
            ("—  Not Evaluated",n,           "595959", "F2F2F2"),
        ], start=4):
            _summary_stat_row(ws, i, lbl, val, fg, bg)
        return

    # Multi-host: per-host table
    hdrs = ["Host / Tab", "Total", "Passed", "Failed", "Not Eval"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color=C_CAT_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=C_CAT_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _BRD
    ws.row_dimensions[4].height = 22

    for ri, (label, checks) in enumerate(host_data, start=5):
        p  = sum(1 for c in checks if c.get("status") == "PASSED")
        f  = sum(1 for c in checks if c.get("status") == "FAILED")
        n  = len(checks) - p - f
        alt = ri % 2 == 0
        bg  = C_ALT if alt else C_WHITE
        for ci, val in enumerate([label, len(checks), p, f, n], start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=9,
                               bold=(ci > 1),
                               color=(C_PASS_FG if ci == 3 and p > 0 else
                                      C_FAIL_FG if ci == 4 and f > 0 else "000000"))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                    vertical="center", indent=1 if ci == 1 else 0)
            c.border    = _BRD
        ws.row_dimensions[ri].height = 18

    # Total row
    tr = 5 + len(host_data)
    all_checks_flat = [c for _, chks in host_data for c in chks]
    tp = sum(1 for c in all_checks_flat if c.get("status") == "PASSED")
    tf = sum(1 for c in all_checks_flat if c.get("status") == "FAILED")
    tn = len(all_checks_flat) - tp - tf
    # Use consolidated checks for total (unique per check not per host)
    consol_label = next((lbl for lbl, _ in host_data if lbl == "Consolidated"), None)
    if consol_label:
        consol_checks = next(chks for lbl, chks in host_data if lbl == "Consolidated")
        tp2 = sum(1 for c in consol_checks if c.get("status") == "PASSED")
        tf2 = sum(1 for c in consol_checks if c.get("status") == "FAILED")
        tn2 = base_count - tp2 - tf2
    else:
        tp2, tf2, tn2 = tp, tf, tn

    for ci, (val, al) in enumerate(
            [("TOTAL (Consolidated)", "left"), (base_count, "center"),
             (tp2, "center"), (tf2, "center"), (tn2, "center")], start=1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font      = Font(name="Arial", bold=True, color=C_HDR_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
        c.alignment = Alignment(horizontal=al, vertical="center",
                                indent=1 if al == "left" else 0)
        c.border    = _BRD
    ws.row_dimensions[tr].height = 22


def _summary_stat_row(ws, row, label, val, fg, bg):
    for ci, (v, al) in enumerate([(label, "left"), (val, "center")], start=1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font      = Font(name="Arial", bold=True, size=10, color=fg)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=al, vertical="center",
                                indent=1 if al == "left" else 0)
        c.border    = _BRD
    ws.row_dimensions[row].height = 22


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(base_checks: list, bench_name: str,
                csv_paths: list, output_path: str):
    auto   = bool(csv_paths)
    multi  = len(csv_paths) > 1
    wb     = Workbook()
    wb.remove(wb.active)

    # ── Parse each CSV ────────────────────────────────────────────────────────
    host_results = []   # [(label, results_dict), ...]
    if auto:
        print(f"\n[2] Parsing {len(csv_paths)} CSV file(s)…")
        for cp in csv_paths:
            label, results = parse_csv_file(cp)
            host_results.append((label, results))
            print(f"    {cp.name:<40} → {len(results)} checks  (host: {label})")

    # ── Build per-host check lists ────────────────────────────────────────────
    host_checks = []   # [(label, filled_checks), ...]
    if auto:
        for label, results in host_results:
            filled, matched = apply_csv_to_checks(base_checks, results)
            host_checks.append((label, filled))
            print(f"    Matched {matched}/{len(base_checks)} for [{label}]")

    # ── Consolidated (multi-host only) ────────────────────────────────────────
    if multi:
        consol = consolidate([chks for _, chks in host_checks], base_checks)
        host_checks.insert(0, ("Consolidated", consol))

    # ── Write sheets ─────────────────────────────────────────────────────────
    print(f"\n[{'3' if auto else '2'}] Building Excel…")

    if not auto:
        # Manual mode — single blank sheet
        blank = deepcopy(base_checks)
        for c in blank:
            c["status"] = ""; c["observed_value"] = ""
        ws = wb.create_sheet(title=_safe_title(bench_name))
        write_checks_sheet(ws, blank, auto_mode=False)
        write_summary_sheet(wb, bench_name, [], len(base_checks))
    else:
        # Deduplicate tab names (e.g. two CSVs from same IP)
        used_tabs: dict = {}
        for label, checks in host_checks:
            safe = _safe_title(label)
            if safe in used_tabs:
                used_tabs[safe] += 1
                safe = _safe_title(f"{label}_{used_tabs[safe]}")
            else:
                used_tabs[safe] = 1
            ws = wb.create_sheet(title=safe)
            write_checks_sheet(ws, checks, auto_mode=True)
            p = sum(1 for c in checks if c.get("status") == "PASSED")
            f = sum(1 for c in checks if c.get("status") == "FAILED")
            print(f"  [tab] {safe:<32}  {len(checks)} checks  "
                  f"{p}✔ {f}✘")

        write_summary_sheet(wb, bench_name, host_checks, len(base_checks))

    wb.save(output_path)

    print(f"\n[✓] Saved  →  {output_path}")
    print(f"    Benchmark  : {bench_name}")
    print(f"    Checks     : {len(base_checks)}")
    print(f"    Mode       : {'Multi-host (%d hosts)' % len(csv_paths) if multi else 'Automatic' if auto else 'Manual'}")
    if multi:
        print(f"    Tabs       : Summary + Consolidated + {len(csv_paths)} host sheet(s)")
    elif auto:
        print(f"    Tabs       : Summary + {host_checks[0][0]}")


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Convert ANY Nessus .audit file to Excel.\n"
                    "Supports CIS Benchmarks, DISA STIGs, and all other formats.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Manual mode
  python audit_to_excel.py --audit CIS_L1.audit

  # Single CSV
  python audit_to_excel.py --audit CIS_L1.audit --csv scan.csv

  # Multiple CSVs (individual files)
  python audit_to_excel.py --audit CIS_L1.audit --csv host1.csv host2.csv host3.csv

  # Folder of CSVs  (all *.csv inside)
  python audit_to_excel.py --audit CIS_L1.audit --csv ./results/

  # Folder + extra file
  python audit_to_excel.py --audit CIS_L1.audit --csv ./results/ extra.csv

  # Multiple audit files + folder
  python audit_to_excel.py --audit L1.audit L2.audit --csv ./results/ -o report.xlsx
""",
    )
    ap.add_argument("--audit", nargs="+", required=True, metavar="FILE")
    ap.add_argument("--csv",   nargs="*", default=[],   metavar="FILE_OR_FOLDER")
    ap.add_argument("--output", "-o", default="audit_report.xlsx")
    args = ap.parse_args()

    # Validate audit files
    missing = [f for f in args.audit if not Path(f).exists()]
    if missing:
        for f in missing:
            print(f"[ERROR] Not found: {f}", file=sys.stderr)
        sys.exit(1)

    # Expand CSV paths (files + folders)
    csv_paths = _expand_csv_paths(args.csv) if args.csv else []
    multi = len(csv_paths) > 1

    print(f"\n{'='*60}")
    if not csv_paths:
        print("  Mode : MANUAL  (audit only)")
    elif multi:
        print(f"  Mode : AUTOMATIC — MULTI-HOST  ({len(csv_paths)} CSV files)")
    else:
        print("  Mode : AUTOMATIC  (single CSV)")
    print(f"{'='*60}")

    print(f"\n[1] Parsing {len(args.audit)} audit file(s)…")
    base_checks, bench_name = parse_audit_files(args.audit)
    print(f"    Total unique checks : {len(base_checks)}")
    print(f"    Benchmark           : {bench_name}")

    if not base_checks:
        print("[ERROR] No checks found.")
        sys.exit(1)

    build_excel(base_checks, bench_name, csv_paths, args.output)


if __name__ == "__main__":
    main()
