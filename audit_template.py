#!/usr/bin/env python3
"""
audit_template.py  —  Phase 1: Template Generator
==================================================
Reads one or more Nessus .audit files and produces a blank Excel template.
Status and Observed Value are left empty — fill them manually or use
report_generator.py to auto-fill from a Nessus CSV export.

Usage:
    python audit_template.py --audit CIS_L1.audit
    python audit_template.py --audit L1.audit L2.audit -o my_template.xlsx

Supported audit formats:
    CIS Benchmarks  (Windows, Linux, macOS, M365 …)
    DISA STIGs      (Oracle, Windows, RHEL …)
    Any Nessus .audit file with <custom_item> or <report> blocks
"""

import re
import sys
import math
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
#  ROW HEIGHT  (openpyxl never auto-sizes — must be calculated)
# ─────────────────────────────────────────────────────────────────────────────

_FONT_PT  = 9
_LINE_H   = _FONT_PT * 1.35
_CHARS_CW = 1.05          # Excel col-width unit ≈ this many chars


def _line_count(text: str, col_w: float) -> float:
    if not text:
        return 1.0
    cpl = max(1, col_w * _CHARS_CW)
    n   = 0.0
    for para in str(text).split("\n"):
        p = para.strip()
        n += 0.4 if not p else math.ceil(len(p) / cpl)
    return max(1.0, n)


def _row_height(pairs: list) -> float:
    """pairs = [(text, col_width), ...] for every wrapped cell in the row."""
    mx = max((_line_count(t, w) for t, w in pairs), default=1.0)
    return max(18, min(mx * _LINE_H + 4, 400))


# ─────────────────────────────────────────────────────────────────────────────
#  AUDIT FILE PARSER
# ─────────────────────────────────────────────────────────────────────────────

_BLOCK_RE = re.compile(
    r'<(custom_item|report[^>]*)>(.*?)</(custom_item|report)>',
    re.DOTALL | re.IGNORECASE,
)
_KV_RE = re.compile(
    r'^\s*([\w]+)\s*:\s*(?:"((?:[^"\\]|\\.)*?)"|([^\n\r]+))',
    re.MULTILINE,
)

# Patterns that identify a real benchmark check ID
_BENCH_PATS = [
    re.compile(r'^\d+\.\d+'),                          # CIS:  1.1.1 ...
    re.compile(r'^[A-Z][A-Z0-9]{1,8}-\d{2}-\d{4,}'),  # DISA: O19C-00-000200
    re.compile(r'^[A-Z]{2,10}-\d{4,}'),
    re.compile(r'^V-\d{4,}'),
]

# @VARIABLE@ → human-readable default values
_VAR_DEFAULTS = {
    "@PASSWORD_HISTORY@":        "24 or more passwords",
    "@MAXIMUM_PASSWORD_AGE@":    "1–365 days",
    "@MINIMUM_PASSWORD_AGE@":    "1 or more day",
    "@MINIMUM_PASSWORD_LENGTH@": "14 or more characters",
    "@LOCKOUT_DURATION@":        "15 or more minutes",
    "@LOCKOUT_THRESHOLD@":       "1–5 invalid attempts",
    "@LOCKOUT_RESET@":           "15 or more minutes",
}


def _is_check(desc: str) -> bool:
    return any(p.match(desc.strip()) for p in _BENCH_PATS)


def _parse_block(block: str) -> dict:
    item = {}
    for m in _KV_RE.finditer(block):
        k = m.group(1).lower().strip()
        v = (m.group(2) if m.group(2) is not None else m.group(3) or "").strip()
        v = v.replace('\\"', '"').replace('\\n', '\n').replace('\\t', '\t')
        if k not in item:
            item[k] = v
    return item


def _clean(text: str) -> str:
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _default_value(value_data: str) -> str:
    """Convert raw value_data from audit file to a human-readable string.
    Returns empty string if nothing meaningful found — never returns 'N/A'.
    """
    if not value_data:
        return ""
    for ph, human in _VAR_DEFAULTS.items():
        if ph in value_data:
            return human
    # Numeric range [n..MAX] or [n..m]
    val = re.sub(
        r'\[(\d+)\.\.(MAX|\d+)\]',
        lambda m: (f"{m.group(1)} or more" if m.group(2) == "MAX"
                   else f"{m.group(1)} to {m.group(2)}"),
        value_data,
    )
    # AUDIT_SET  "x" || "y"
    val = val.replace(' || ', ' or ')
    # Strip outer quotes and common boilerplate
    val = val.strip('"').strip()
    # If it's just a bare regex pattern or very long, trim to something readable
    if len(val) > 150:
        val = val[:150] + "…"
    return val


def _bench_name_from_file(text: str, filepath: str) -> str:
    m = re.search(r'<display_name>(.*?)</display_name>', text, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r'#\s*description\s*:\s*This .audit is designed against the (.+)', text)
    if m:
        return m.group(1).strip()
    return Path(filepath).stem.replace('_', ' ')


def _sort_key(benchmark: str):
    m = re.match(r'^([\d.]+)', benchmark.strip())
    if m:
        return [int(p) if p.isdigit() else p for p in m.group(1).split('.')]
    m2 = re.match(r'^([A-Z0-9-]+)', benchmark.strip())
    return [m2.group(1)] if m2 else [benchmark]


def parse_audit(filepaths: list) -> tuple:
    """Returns (checks, benchmark_name).
    checks = list of dicts: benchmark, description, default_value, recommendation
    """
    checks    = []
    seen      = set()
    bench_name = ""

    for fp in filepaths:
        text = Path(fp).read_text(encoding="utf-8", errors="replace")
        if not bench_name:
            bench_name = _bench_name_from_file(text, fp)

        cnt = 0
        for m in _BLOCK_RE.finditer(text):
            item = _parse_block(m.group(2))
            desc = item.get("description", "").strip()
            if not desc or not _is_check(desc) or desc in seen:
                continue
            seen.add(desc)
            checks.append({
                "benchmark":      desc,
                "description":    _clean(item.get("info", "")),
                "default_value":  _default_value(item.get("value_data", "")),
                "recommendation": _clean(item.get("solution", "")),
            })
            cnt += 1
        print(f"  {Path(fp).name}: {cnt} checks")

    checks.sort(key=lambda c: _sort_key(c["benchmark"]))
    return checks, bench_name or "Checks"


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL STYLES
# ─────────────────────────────────────────────────────────────────────────────

C_HDR_BG = "1F3864";  C_HDR_FG = "FFFFFF"
C_ALT    = "EBF3FB";  C_WHITE  = "FFFFFF"
C_BORDER = "8EA9C1"

_S   = Side(style="thin", color=C_BORDER)
_BRD = Border(left=_S, right=_S, top=_S, bottom=_S)

# (header, width, wrap, align)
COLUMNS = [
    ("S.NO",           7,   False, "center"),
    ("Benchmark",      52,  True,  "left"),
    ("Description",    70,  True,  "left"),
    ("Status",         14,  False, "center"),
    ("Default Value",  28,  True,  "left"),
    ("Observed Value", 32,  True,  "left"),
    ("Recommendation", 62,  True,  "left"),
]


def _hdr(cell, text: str):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color=C_HDR_FG, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _BRD


def _dat(cell, value, bg: str, align: str, wrap: bool):
    cell.value     = value
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", size=_FONT_PT, color="000000")
    cell.alignment = Alignment(horizontal=align, vertical="top",
                               wrap_text=wrap)
    cell.border    = _BRD


def write_sheet(ws, checks: list):
    ws.row_dimensions[1].height = 30
    for ci, (label, width, _, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        _hdr(ws.cell(row=1, column=ci), label)
    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    for ri, chk in enumerate(checks, start=1):
        row = ri + 1
        bg  = C_ALT if ri % 2 == 0 else C_WHITE

        vals = [
            (ri,                       "center", False),
            (chk["benchmark"],         "left",   True),
            (chk["description"],       "left",   True),
            ("",                       "center", False),   # Status — blank
            (chk["default_value"],     "left",   True),
            ("",                       "left",   True),    # Observed Value — blank
            (chk["recommendation"],    "left",   True),
        ]
        for ci, (val, align, wrap) in enumerate(vals, start=1):
            _dat(ws.cell(row=row, column=ci),
                 val if ci > 1 else ri, bg, align, wrap)

        wrapped = [(str(v), COLUMNS[ci-1][1])
                   for ci, (v, _, wrap) in enumerate(vals, start=1) if wrap]
        ws.row_dimensions[row].height = _row_height(wrapped)


def write_summary(wb, bench_name: str, total: int):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14

    ws.merge_cells("A1:B1")
    t           = ws.cell(row=1, column=1, value=bench_name)
    t.font      = Font(name="Arial", bold=True, size=13, color=C_HDR_FG)
    t.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    sub = ws.cell(row=2, column=1,
                  value="Template — fill Status and Observed Value manually "
                        "or use report_generator.py")
    sub.font      = Font(name="Arial", size=9, italic=True, color="595959")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    for ci, (v, al) in enumerate([("Total Checks", "left"),
                                   (total,          "center")], start=1):
        c           = ws.cell(row=4, column=ci, value=v)
        c.font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FG)
        c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
        c.alignment = Alignment(horizontal=al, vertical="center",
                                indent=1 if al == "left" else 0)
        c.border    = _BRD
    ws.row_dimensions[4].height = 24


def build_template(checks: list, bench_name: str, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    safe = re.sub(r'[:/\\?\[\]*]', '-', bench_name)[:31] or "Checks"
    ws   = wb.create_sheet(title=safe)
    write_sheet(ws, checks)
    write_summary(wb, bench_name, len(checks))

    wb.save(output_path)
    print(f"\n[✓] Template saved → {output_path}")
    print(f"    Benchmark : {bench_name}")
    print(f"    Checks    : {len(checks)}")
    print(f"    Fill Status + Observed Value manually, or run report_generator.py")


# ─────────────────────────────────────────────────────────────────────────────
#  CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Phase 1 — Generate a blank Excel template from a Nessus .audit file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python audit_template.py --audit CIS_L1.audit
  python audit_template.py --audit L1.audit L2.audit -o template.xlsx
  python audit_template.py --audit DISA_Oracle.audit -o oracle_template.xlsx
""")
    ap.add_argument("--audit", nargs="+", required=True, metavar="FILE",
                    help=".audit file(s)")
    ap.add_argument("-o", "--output", default="audit_template.xlsx",
                    help="Output filename  [default: audit_template.xlsx]")
    args = ap.parse_args()

    missing = [f for f in args.audit if not Path(f).exists()]
    if missing:
        for f in missing:
            print(f"[ERROR] Not found: {f}", file=sys.stderr)
        sys.exit(1)

    print(f"[*] Parsing {len(args.audit)} audit file(s)…")
    checks, bench_name = parse_audit(args.audit)
    print(f"    Total checks : {len(checks)}")
    print(f"    Benchmark    : {bench_name}")

    if not checks:
        print("[ERROR] No checks found.")
        sys.exit(1)

    build_template(checks, bench_name, args.output)


if __name__ == "__main__":
    main()
