#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║           Nessus CIS Audit → Excel  (Manual + Automatic modes)             ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                              ║
║  MANUAL MODE  — audit file only, Status/Observed Value left blank           ║
║    python nessus_audit_to_excel.py --audit CIS_L1.audit                     ║
║                                                                              ║
║  AUTOMATIC MODE  — audit + Nessus CSV export, fully populated               ║
║    python nessus_audit_to_excel.py --audit CIS_L1.audit --csv scan.csv      ║
║                                                                              ║
║  Multiple audit files (L1 + L2):                                            ║
║    python nessus_audit_to_excel.py --audit L1.audit L2.audit --csv scan.csv ║
║                                                                              ║
║  Custom output:                                                              ║
║    python nessus_audit_to_excel.py --audit L1.audit --csv s.csv -o rep.xlsx ║
╚══════════════════════════════════════════════════════════════════════════════╝

Columns:
  S.NO | Benchmark | Description | Status | Default Value | Observed Value |
  Observation | Recommendation

In Automatic mode, Status and Observed Value are filled from the CSV.
In Manual mode, both columns are blank — fill them after your manual check.
"""

import re
import sys
import csv
import math
import argparse
from pathlib import Path
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  ROW HEIGHT CALCULATOR
#  openpyxl does NOT auto-size rows with wrap_text.
#  We calculate the required height from text length + column width.
# ══════════════════════════════════════════════════════════════════════════════

FONT_PT       = 9
LINE_HEIGHT   = FONT_PT * 1.35          # pt per line at Arial 9
MIN_ROW_H     = 18
MAX_ROW_H     = 400
CHARS_PER_UNIT = 1.05                   # Excel col-width unit → chars


def _lines(text: str, col_width: float) -> float:
    if not text:
        return 1.0
    cpl = max(1, col_width * CHARS_PER_UNIT)
    tot = 0.0
    for para in str(text).split("\n"):
        p = para.strip()
        tot += 0.4 if not p else math.ceil(len(p) / cpl)
    return max(1.0, tot)


def row_height(cell_pairs: list) -> float:
    """cell_pairs = [(text, col_width), ...] for every wrapped cell in the row."""
    mx = max((_lines(t, w) for t, w in cell_pairs), default=1.0)
    return max(MIN_ROW_H, min(mx * LINE_HEIGHT + 4, MAX_ROW_H))


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION → TAB MAPPING
# ══════════════════════════════════════════════════════════════════════════════

SECTION_MAP = OrderedDict([
    ("1.1",   ("Password Policy",        "1.1  Password Policy")),
    ("1.2",   ("Account Lockout",        "1.2  Account Lockout Policy")),
    ("2.2",   ("User Rights",            "2.2  User Rights Assignment")),
    ("2.3",   ("Security Options",       "2.3  Security Options")),
    ("9.1",   ("Firewall - Domain",      "9.1  Windows Firewall - Domain Profile")),
    ("9.2",   ("Firewall - Private",     "9.2  Windows Firewall - Private Profile")),
    ("9.3",   ("Firewall - Public",      "9.3  Windows Firewall - Public Profile")),
    ("17.1",  ("Audit - Acct Logon",     "17.1 Advanced Audit - Account Logon")),
    ("17.2",  ("Audit - Acct Mgmt",      "17.2 Advanced Audit - Account Management")),
    ("17.3",  ("Audit - Detail Track",   "17.3 Advanced Audit - Detailed Tracking")),
    ("17.5",  ("Audit - Logon-Logoff",   "17.5 Advanced Audit - Logon / Logoff")),
    ("17.6",  ("Audit - Object Access",  "17.6 Advanced Audit - Object Access")),
    ("17.7",  ("Audit - Policy Change",  "17.7 Advanced Audit - Policy Change")),
    ("17.8",  ("Audit - Privilege Use",  "17.8 Advanced Audit - Privilege Use")),
    ("17.9",  ("Audit - System",         "17.9 Advanced Audit - System")),
    ("18.1",  ("AT - Control Panel",     "18.1  Admin Templates - Control Panel")),
    ("18.4",  ("AT - MSS Legacy",        "18.4  Admin Templates - MSS (Legacy)")),
    ("18.5",  ("AT - Network",           "18.5  Admin Templates - Network")),
    ("18.6",  ("AT - Printers",          "18.6  Admin Templates - Printers")),
    ("18.7",  ("AT - Start Menu",        "18.7  Admin Templates - Start Menu")),
    ("18.9",  ("AT - System",            "18.9  Admin Templates - System")),
    ("18.10", ("AT - Windows Comp",      "18.10 Admin Templates - Windows Components")),
    ("19.5",  ("User AT - IE",           "19.5  User Admin Templates - Internet Explorer")),
    ("19.7",  ("User AT - Win Comp",     "19.7  User Admin Templates - Windows Components")),
])
FALLBACK_TAB = ("Other Checks", "Other Checks")


def get_category(bench_num: str):
    for prefix, names in SECTION_MAP.items():
        if bench_num == prefix or bench_num.startswith(prefix + "."):
            return names
    return FALLBACK_TAB


# ══════════════════════════════════════════════════════════════════════════════
#  AUDIT FILE PARSER
# ══════════════════════════════════════════════════════════════════════════════

_BLOCK_RE = re.compile(r'<custom_item>(.*?)</custom_item>', re.DOTALL | re.IGNORECASE)
_KV_RE    = re.compile(
    r'^\s*([\w]+)\s*:\s*(?:"((?:[^"\\]|\\.)*?)"|([^\n\r]+))',
    re.MULTILINE,
)

_VAR_DEFAULTS = {
    "@PASSWORD_HISTORY@":       "24 or more passwords",
    "@MAXIMUM_PASSWORD_AGE@":   "1 to 365 days",
    "@MINIMUM_PASSWORD_AGE@":   "1 or more day(s)",
    "@MINIMUM_PASSWORD_LENGTH@":"14 or more characters",
    "@LOCKOUT_DURATION@":       "15 or more minutes",
    "@LOCKOUT_THRESHOLD@":      "1 to 5 invalid attempts",
    "@LOCKOUT_RESET@":          "15 or more minutes",
    "@LAPS_PASSWORD_LENGTH@":   "15 or more characters",
    "@LAPS_PASSWORD_AGE@":      "Up to 30 days",
    "@LAPS_GRACE_PERIOD@":      "1 to 8 hours",
    "@PASSWORD_AGE_PROMPT@":    "5 to 14 days",
    "@LEGAL_NOTICE_TEXT@":      "Banner text (site-specific)",
    "@LEGAL_CAPTION_TEXT@":     "Banner caption (site-specific)",
}


def _parse_block(block: str) -> dict:
    item = {}
    for m in _KV_RE.finditer(block):
        key = m.group(1).lower().strip()
        val = m.group(2) if m.group(2) is not None else (m.group(3) or "")
        val = val.strip().replace('\\"', '"')
        if key not in item:
            item[key] = val
    return item


def _clean(text: str) -> str:
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _default_val(value_data: str) -> str:
    if not value_data:
        return ""
    for ph, human in _VAR_DEFAULTS.items():
        if ph in value_data:
            return human
    val = re.sub(
        r'\[(\d+)\.\.(MAX|\d+)\]',
        lambda m: (f"{m.group(1)} or more"
                   if m.group(2) == "MAX"
                   else f"{m.group(1)} to {m.group(2)}"),
        value_data,
    )
    return val.replace(' || ', ' or ').strip('"').strip()


def _bench_num(description: str) -> str:
    m = re.match(r'^([\d]+(?:\.[\d]+)*)', description.strip())
    return m.group(1) if m else ""


def _detect_level(filepath: str, text: str) -> str:
    fn = Path(filepath).name.upper()
    return "L2" if ("L2" in fn or "LEVEL_2" in fn or "LEVEL|2" in text) else "L1"


def parse_audit_files(filepaths: list) -> list:
    checks, seen = [], set()
    for fp in filepaths:
        text  = Path(fp).read_text(encoding="utf-8", errors="replace")
        level = _detect_level(fp, text)
        count = 0
        for m in _BLOCK_RE.finditer(text):
            item     = _parse_block(m.group(1))
            desc_raw = item.get("description", "").strip()
            if not re.match(r'^\d+\.', desc_raw) or desc_raw in seen:
                continue
            seen.add(desc_raw)
            bn              = _bench_num(desc_raw)
            tab_name, label = get_category(bn)
            checks.append({
                "bench_num":      bn,
                "benchmark":      desc_raw,
                "description":    _clean(item.get("info", "")),
                "default_value":  _default_val(item.get("value_data", "")),
                "recommendation": _clean(item.get("solution", "")),
                "tab_name":       tab_name,
                "category_label": label,
                "level":          level,
                # filled later by CSV merge
                "status":         "",
                "observed_value": "",
                "host":           "",
            })
            count += 1
        print(f"    {Path(fp).name}: {count} checks (level {level})")

    checks.sort(key=lambda c: [int(p) if p.isdigit() else p
                                for p in re.split(r'\.', c["bench_num"])])
    return checks


# ══════════════════════════════════════════════════════════════════════════════
#  NESSUS CSV PARSER  &  MERGER
# ══════════════════════════════════════════════════════════════════════════════

# Possible column name variants across Nessus versions
_NAME_COLS    = ("Name", "Plugin Name", "Check Name", "name", "plugin_name")
_STATUS_COLS  = ("Risk", "Status", "Result", "risk", "status", "result")
_OUTPUT_COLS  = ("Plugin Output", "Output", "Actual Value",
                 "plugin_output", "output", "actual_value")
_HOST_COLS    = ("Host", "IP Address", "host", "ip_address")


def _find_col(headers: list, candidates: tuple) -> str:
    """Return the first candidate column name that exists in headers."""
    for c in candidates:
        if c in headers:
            return c
    return ""


def _norm_name(s: str) -> str:
    """Normalise a benchmark name to just its leading number for matching."""
    m = re.match(r'^([\d]+(?:\.[\d]+)*)', s.strip())
    return m.group(1) if m else re.sub(r'[^a-z0-9]', '', s.lower())


def _extract_actual(output: str) -> str:
    """
    Pull the actual/observed value from Plugin Output text.
    Nessus formats this as:
        'Actual Value: <value>\nPolicy Value: <value>'
    or simply raw output text.
    """
    if not output:
        return ""
    # Try explicit "Actual Value:" label first
    m = re.search(r'Actual\s+Value\s*:\s*(.+?)(?:\n|Policy|$)', output,
                  re.IGNORECASE | re.DOTALL)
    if m:
        return m.group(1).strip()
    # Fall back: first meaningful line
    for line in output.strip().splitlines():
        line = line.strip()
        if line and not line.lower().startswith("policy"):
            return line
    return output.strip()[:200]


def _norm_status(raw: str) -> str:
    r = raw.strip().upper()
    if r in ("PASSED", "PASS"):    return "PASSED"
    if r in ("FAILED", "FAIL"):    return "FAILED"
    if r in ("WARNING", "WARN"):   return "WARNING"
    if r in ("ERROR",):            return "ERROR"
    return r or "INFO"


def parse_csv_files(filepaths: list) -> dict:
    """
    Parse one or more Nessus compliance CSV exports.
    Returns { bench_number_prefix → {status, observed_value, host} }
    Multiple rows for the same check (multiple hosts) are stored as lists.
    """
    results: dict = {}   # bench_num → list of dicts

    for fp in filepaths:
        print(f"    {Path(fp).name}: ", end="", flush=True)
        count = 0
        try:
            with open(fp, newline="", encoding="utf-8-sig", errors="replace") as f:
                # Sniff delimiter — could be comma or tab
                sample = f.read(4096); f.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
                reader  = csv.DictReader(f, dialect=dialect)

                name_col   = _find_col(reader.fieldnames or [], _NAME_COLS)
                status_col = _find_col(reader.fieldnames or [], _STATUS_COLS)
                output_col = _find_col(reader.fieldnames or [], _OUTPUT_COLS)
                host_col   = _find_col(reader.fieldnames or [], _HOST_COLS)

                if not name_col or not status_col:
                    print(f"⚠  Cannot identify Name/Status columns "
                          f"(found: {reader.fieldnames})")
                    continue

                for row in reader:
                    name   = row.get(name_col, "").strip()
                    status = _norm_status(row.get(status_col, ""))
                    output = _extract_actual(row.get(output_col, ""))
                    host   = row.get(host_col, "").strip()

                    if not name:
                        continue

                    key = _norm_name(name)
                    if key not in results:
                        results[key] = []
                    results[key].append({
                        "status":         status,
                        "observed_value": output,
                        "host":           host,
                        "raw_name":       name,
                    })
                    count += 1

        except Exception as e:
            print(f"⚠  Error reading CSV: {e}")
            continue

        print(f"{count} rows loaded")

    return results


def _merge(checks: list, csv_results: dict) -> list:
    """
    Merge CSV results into checks list.
    Matching strategy:
      1. Exact benchmark number prefix match  (most reliable)
      2. Substring match on normalised name   (fallback)
    When multiple hosts reported the same check, we show:
      - FAILED if ANY host failed (worst-case)
      - observed_value = all unique actual values joined
      - host = all unique hosts joined
    """
    matched = 0
    for chk in checks:
        key = chk["bench_num"]              # e.g. "1.1.1"
        rows = csv_results.get(key, [])

        if not rows:
            # Try suffix match for keys like "1.1.1" inside "1.1.1.1" etc.
            rows = next(
                (v for k, v in csv_results.items()
                 if k == key or k.startswith(key + ".") or key.startswith(k + ".")),
                [],
            )

        if not rows:
            continue

        # Aggregate multi-host results
        statuses  = [r["status"] for r in rows]
        obs_vals  = list(dict.fromkeys(r["observed_value"] for r in rows if r["observed_value"]))
        hosts     = list(dict.fromkeys(r["host"] for r in rows if r["host"]))

        # Worst-case status
        if "FAILED"  in statuses: agg_status = "FAILED"
        elif "ERROR" in statuses: agg_status = "ERROR"
        elif "WARNING" in statuses: agg_status = "WARNING"
        elif all(s == "PASSED" for s in statuses): agg_status = "PASSED"
        else: agg_status = statuses[0]

        chk["status"]         = agg_status
        chk["observed_value"] = "\n".join(obs_vals)
        chk["host"]           = ", ".join(hosts)
        matched += 1

    print(f"\n    Matched {matched} / {len(checks)} checks to CSV results")
    return checks


# ══════════════════════════════════════════════════════════════════════════════
#  COLOUR PALETTE
# ══════════════════════════════════════════════════════════════════════════════

C_HDR_BG   = "1F3864"   # deep navy
C_HDR_FG   = "FFFFFF"
C_CAT_BG   = "2E75B6"   # CIS blue (category header rows)
C_CAT_FG   = "FFFFFF"
C_PASS_ROW = "EBF5E1"   # very light green  (row fill for PASSED)
C_PASS_FG  = "375623"   # dark green text
C_FAIL_ROW = "FDE9E8"   # very light red    (row fill for FAILED)
C_FAIL_FG  = "9C0006"   # dark red text
C_WARN_ROW = "FFF3CD"   # light amber       (row fill for WARNING)
C_WARN_FG  = "7F5200"
C_ALT_ROW  = "EBF3FB"   # light blue-grey   (no-CSV alternating rows)
C_WHITE    = "FFFFFF"
C_BORDER   = "8EA9C1"

_S   = Side(style="thin", color=C_BORDER)
_BRD = Border(left=_S, right=_S, top=_S, bottom=_S)


def _status_colours(status: str):
    """Return (row_bg, status_fg) for a given status string."""
    s = status.upper()
    if s == "PASSED":  return C_PASS_ROW, C_PASS_FG
    if s == "FAILED":  return C_FAIL_ROW, C_FAIL_FG
    if s == "WARNING": return C_WARN_ROW, C_WARN_FG
    return C_WHITE, "595959"


# ══════════════════════════════════════════════════════════════════════════════
#  COLUMN DEFINITIONS
#  (label, col_width, wrap_text, h-align)
# ══════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    ("S.NO",           7,   False, "center"),
    ("Benchmark",      50,  True,  "left"),
    ("Description",    68,  True,  "left"),
    ("Status",         14,  False, "center"),
    ("Default Value",  26,  True,  "left"),
    ("Observed Value", 28,  True,  "left"),
    ("Observation",    30,  True,  "left"),
    ("Recommendation", 60,  True,  "left"),
]


def _hdr(cell, text: str):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color=C_HDR_FG, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _BRD


def _data(cell, value, bg: str, fg: str, align: str, wrap: bool, bold=False):
    cell.value     = value
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", size=FONT_PT, color=fg, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border    = _BRD


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET WRITER
# ══════════════════════════════════════════════════════════════════════════════

def _write_category_sheet(ws, checks: list, auto_mode: bool):
    # ── Column widths + header ─────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    for ci, (label, width, wrap, align) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        _hdr(ws.cell(row=1, column=ci), label)

    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Data rows ──────────────────────────────────────────────────────────
    for ri, chk in enumerate(checks, start=1):
        excel_row = ri + 1
        status    = chk["status"]

        if auto_mode and status:
            row_bg, status_fg = _status_colours(status)
        else:
            row_bg    = C_ALT_ROW if ri % 2 == 0 else C_WHITE
            status_fg = "595959"

        row_values = [
            (ri,                        "center", False, False),
            (chk["benchmark"],          "left",   True,  False),
            (chk["description"],        "left",   True,  False),
            (status,                    "center", False, True),   # Status — bold
            (chk["default_value"],      "left",   True,  False),
            (chk["observed_value"],     "left",   True,  False),
            ("",                        "left",   True,  False),  # Observation blank
            (chk["recommendation"],     "left",   True,  False),
        ]

        for ci, ((value, align, wrap, bold), (_, _, _, _)) in enumerate(
                zip(row_values, COLUMNS), start=1):
            label = COLUMNS[ci - 1][0]
            # Status cell gets its own fg colour regardless of row
            fg = status_fg if label == "Status" else "000000"
            _data(ws.cell(row=excel_row, column=ci),
                  value, row_bg, fg, align, wrap, bold)

        # ── Calculated row height ──────────────────────────────────────────
        wrapped = [(str(val), COLUMNS[ci - 1][1])
                   for ci, (val, _, wrap, _) in enumerate(row_values, start=1)
                   if wrap]
        ws.row_dimensions[excel_row].height = row_height(wrapped)


# ══════════════════════════════════════════════════════════════════════════════
#  SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════════

def _write_summary(wb: Workbook, checks: list, tab_order: list,
                   auto_mode: bool, hosts: list):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16

    # ── Title ──────────────────────────────────────────────────────────────
    mode_label = "Automatic Mode (CSV merged)" if auto_mode else "Manual Mode"
    ws.merge_cells("A1:E1")
    t           = ws.cell(row=1, column=1,
                          value="CIS Windows Server 2025 – Hardening Audit Report")
    t.font      = Font(name="Arial", bold=True, size=13, color=C_HDR_FG)
    t.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    sub = ws.cell(row=2, column=1,
                  value=f"Benchmark v1.0.0  •  L1 MS  •  {mode_label}"
                        + (f"  •  Host(s): {', '.join(hosts)}" if hosts else ""))
    sub.font      = Font(name="Arial", size=9, italic=True, color="595959")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Overall counts (auto mode only) ───────────────────────────────────
    if auto_mode:
        total   = len(checks)
        passed  = sum(1 for c in checks if c["status"] == "PASSED")
        failed  = sum(1 for c in checks if c["status"] == "FAILED")
        warning = sum(1 for c in checks if c["status"] == "WARNING")
        no_eval = total - passed - failed - warning

        stat_rows = [
            ("Total Checks",       total,   C_HDR_BG,    C_HDR_FG),
            ("✔  Passed",          passed,  "375623",    C_PASS_ROW),
            ("✘  Failed",          failed,  "9C0006",    C_FAIL_ROW),
            ("⚠  Warning",         warning, "7F5200",    C_WARN_ROW),
            ("—  Not Evaluated",   no_eval, "595959",    "F2F2F2"),
        ]
        for row_i, (label, val, fg, bg) in enumerate(stat_rows, start=4):
            for ci, (v, al) in enumerate([(label, "left"), (val, "center")], start=1):
                c           = ws.cell(row=row_i, column=ci, value=v)
                c.font      = Font(name="Arial", bold=True, size=10, color=fg)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal=al, vertical="center",
                                        indent=1 if al == "left" else 0)
                c.border    = _BRD
            ws.merge_cells(f"B{row_i}:E{row_i}")
            ws.row_dimensions[row_i].height = 22
        hdr_row = 10
    else:
        hdr_row = 4

    # ── Per-category table ─────────────────────────────────────────────────
    hdr_cols = (["Category", "Checks", "Passed", "Failed", "Not Evaluated"]
                if auto_mode else ["Category", "Checks"])

    for ci, h in enumerate(hdr_cols, start=1):
        c           = ws.cell(row=hdr_row, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color=C_CAT_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=C_CAT_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _BRD
    ws.row_dimensions[hdr_row].height = 22

    cat_data: dict = {}
    for chk in checks:
        t = chk["tab_name"]
        cat_data.setdefault(t, {"label": chk["category_label"],
                                "total": 0, "passed": 0,
                                "failed": 0, "no_eval": 0})
        cat_data[t]["total"] += 1
        s = chk["status"].upper()
        if s == "PASSED":  cat_data[t]["passed"]  += 1
        elif s == "FAILED": cat_data[t]["failed"]  += 1
        else:               cat_data[t]["no_eval"] += 1

    for ri, tab in enumerate(tab_order, start=hdr_row + 1):
        d   = cat_data.get(tab, {})
        alt = ri % 2 == 0
        bg  = C_ALT_ROW if alt else C_WHITE

        row_vals = [d.get("label", tab), d.get("total", 0)]
        if auto_mode:
            row_vals += [d.get("passed", 0), d.get("failed", 0), d.get("no_eval", 0)]

        for ci, v in enumerate(row_vals, start=1):
            c           = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(name="Arial", size=9,
                               bold=(ci > 1),
                               color=(C_PASS_FG if (ci == 3 and auto_mode and v > 0)
                                      else C_FAIL_FG if (ci == 4 and auto_mode and v > 0)
                                      else "000000"))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                    vertical="center",
                                    indent=1 if ci == 1 else 0)
            c.border    = _BRD
        ws.row_dimensions[ri].height = 18

    # Total row
    tr = hdr_row + len(tab_order) + 1
    total_vals = [("Total", "left"),
                  (len(checks), "center")]
    if auto_mode:
        passed_t  = sum(1 for c in checks if c["status"] == "PASSED")
        failed_t  = sum(1 for c in checks if c["status"] == "FAILED")
        no_eval_t = len(checks) - passed_t - failed_t
        total_vals += [(passed_t, "center"), (failed_t, "center"), (no_eval_t, "center")]
    for ci, (v, al) in enumerate(total_vals, start=1):
        c           = ws.cell(row=tr, column=ci, value=v)
        c.font      = Font(name="Arial", bold=True, color=C_HDR_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
        c.alignment = Alignment(horizontal=al, vertical="center",
                                indent=1 if al == "left" else 0)
        c.border    = _BRD
    ws.row_dimensions[tr].height = 20


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(checks: list, output_path: str, auto_mode: bool):
    wb    = Workbook()
    wb.remove(wb.active)

    groups: OrderedDict = OrderedDict()
    for chk in checks:
        groups.setdefault(chk["tab_name"], []).append(chk)

    for tab_name, tab_checks in groups.items():
        safe = re.sub(r'[:/\\?\[\]*]', '-', tab_name)[:31]
        ws   = wb.create_sheet(title=safe)
        _write_category_sheet(ws, tab_checks, auto_mode)

        # Tab colour by worst status in automatic mode
        if auto_mode:
            statuses = {c["status"].upper() for c in tab_checks if c["status"]}
            if "FAILED"  in statuses: ws.sheet_properties.tabColor = "C2212E"
            elif "WARNING" in statuses: ws.sheet_properties.tabColor = "F18C43"
            elif "PASSED" in statuses: ws.sheet_properties.tabColor = "527421"

        p = sum(1 for c in tab_checks if c["status"] == "PASSED")
        f = sum(1 for c in tab_checks if c["status"] == "FAILED")
        n = len(tab_checks) - p - f
        status_str = (f"  {p}✔ {f}✘ {n}?" if auto_mode else "")
        print(f"  [tab] {tab_name:<28}  {len(tab_checks):>3} checks{status_str}")

    # Collect all unique hosts from CSV results
    hosts = list(dict.fromkeys(c["host"] for c in checks if c.get("host")))
    _write_summary(wb, checks, list(groups.keys()), auto_mode, hosts)

    wb.save(output_path)
    print(f"\n[✓] Saved  →  {output_path}")
    print(f"    Mode: {'Automatic (audit + CSV merged)' if auto_mode else 'Manual (audit only)'}")
    print(f"    Tabs: Summary + {len(groups)} category sheets")
    if auto_mode:
        p = sum(1 for c in checks if c["status"] == "PASSED")
        f = sum(1 for c in checks if c["status"] == "FAILED")
        n = len(checks) - p - f
        print(f"    Results: {p} PASSED  |  {f} FAILED  |  {n} Not Evaluated")


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Convert Nessus CIS .audit files to formatted Excel.\n"
                    "Add --csv to auto-populate Pass/Fail from a Nessus scan CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Manual mode — blank Status/Observed Value columns
  python nessus_audit_to_excel.py --audit CIS_L1.audit

  # Automatic mode — merge audit + Nessus CSV export
  python nessus_audit_to_excel.py --audit CIS_L1.audit --csv scan_results.csv

  # L1 + L2 with scan results
  python nessus_audit_to_excel.py --audit L1.audit L2.audit --csv scan.csv

  # Multiple CSV files (e.g. multiple scan targets)
  python nessus_audit_to_excel.py --audit L1.audit --csv host1.csv host2.csv
""",
    )
    ap.add_argument("--audit", nargs="+", required=True, metavar="FILE",
                    help=".audit file(s) — CIS benchmark definitions")
    ap.add_argument("--csv",   nargs="*", default=[], metavar="FILE",
                    help="Nessus compliance CSV export(s) — enables Automatic mode")
    ap.add_argument("--output", "-o", default="cis_hardening_report.xlsx",
                    help="Output Excel filename  [default: cis_hardening_report.xlsx]")
    args = ap.parse_args()

    missing = [f for f in args.audit + (args.csv or []) if not Path(f).exists()]
    if missing:
        for f in missing:
            print(f"[ERROR] File not found: {f}", file=sys.stderr)
        sys.exit(1)

    auto_mode = bool(args.csv)

    print(f"\n{'='*60}")
    print(f"  Mode: {'AUTOMATIC  (audit + CSV merge)' if auto_mode else 'MANUAL  (audit only)'}")
    print(f"{'='*60}")

    # 1. Parse audit files
    print(f"\n[1] Parsing {len(args.audit)} audit file(s)…")
    checks = parse_audit_files(args.audit)
    print(f"    Total unique checks: {len(checks)}")

    # 2. Parse CSV + merge (automatic mode only)
    if auto_mode:
        print(f"\n[2] Parsing {len(args.csv)} CSV file(s)…")
        csv_results = parse_csv_files(args.csv)
        print(f"    Total CSV entries: {sum(len(v) for v in csv_results.values())}")
        print(f"\n[3] Merging…")
        checks = _merge(checks, csv_results)
    else:
        print("\n    No CSV provided — Status and Observed Value will be blank.")
        print("    Tip: add --csv <scan.csv> to auto-populate from a Nessus scan.")

    if not checks:
        print("[ERROR] No checks found.")
        sys.exit(1)

    # 3. Build Excel
    step = 4 if auto_mode else 2
    print(f"\n[{step}] Building Excel workbook…")
    build_excel(checks, args.output, auto_mode)


if __name__ == "__main__":
    main()
